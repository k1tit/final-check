# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import os
import re
import sys
from pathlib import Path
from typing import Optional, Set, Tuple

import numpy as np
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")


def _project_root() -> Path:
    return Path(__file__).resolve().parent


_ROOT = _project_root()
RUNTIME_PATHS_JSON = _ROOT / "runtime_paths.json"

ZERO_FILES_SUBDIR = "1 Нулевые файлы выгрузки макроса + файл исключений"
DATA_DIR = _ROOT / "data"


def _normalize_path_text(raw: object) -> str:
    txt = str(raw or "").strip().strip('"').strip("'")
    if not txt:
        return ""
    return str(Path(os.path.expandvars(os.path.expanduser(txt))))


def _resolve_config_path(raw: object) -> Path:
    txt = _normalize_path_text(raw)
    if not txt:
        return Path()
    p = Path(txt)
    if not p.is_absolute():
        p = _ROOT / p
    return p.resolve()


def _path_for_config(p: Path) -> str:
    p = p.resolve()
    try:
        return p.relative_to(_ROOT.resolve()).as_posix()
    except ValueError:
        return str(p)


def _paths_from_data_dir(data_dir: Path) -> dict[str, Path]:
    return {
        "data_dir": data_dir,
        "base_dir": data_dir / ZERO_FILES_SUBDIR,
        "output_dir": data_dir / "result",
        "exception_file": data_dir / "Exception.xlsx",
    }


def load_runtime_paths_dict() -> dict[str, Path]:
    """Все рабочие пути. Достаточно задать data_dir — остальное выводится автоматически."""
    result = _paths_from_data_dir(DATA_DIR)
    try:
        if RUNTIME_PATHS_JSON.is_file():
            cfg = json.loads(RUNTIME_PATHS_JSON.read_text(encoding="utf-8"))
            if cfg.get("data_dir"):
                result = _paths_from_data_dir(_resolve_config_path(cfg["data_dir"]))
            else:
                if cfg.get("base_dir"):
                    result["base_dir"] = _resolve_config_path(cfg["base_dir"])
                if cfg.get("output_dir"):
                    result["output_dir"] = _resolve_config_path(cfg["output_dir"])
                if cfg.get("exception_file"):
                    result["exception_file"] = _resolve_config_path(cfg["exception_file"])
                if result["base_dir"].name == ZERO_FILES_SUBDIR:
                    result["data_dir"] = result["base_dir"].parent
                elif cfg.get("base_dir"):
                    result["data_dir"] = result["base_dir"].parent
    except Exception:
        pass

    if os.environ.get("REPORTS_DATA_DIR"):
        result = _paths_from_data_dir(_resolve_config_path(os.environ["REPORTS_DATA_DIR"]))
    if os.environ.get("REPORTS_BASE_DIR"):
        result["base_dir"] = _resolve_config_path(os.environ["REPORTS_BASE_DIR"])
    if os.environ.get("REPORTS_OUTPUT_DIR"):
        result["output_dir"] = _resolve_config_path(os.environ["REPORTS_OUTPUT_DIR"])
    if os.environ.get("REPORTS_EXCEPTION_FILE"):
        result["exception_file"] = _resolve_config_path(os.environ["REPORTS_EXCEPTION_FILE"])
    if result["base_dir"].name == ZERO_FILES_SUBDIR:
        result["data_dir"] = result["base_dir"].parent
    return result


def _load_runtime_paths() -> tuple[Path, Path, Path]:
    d = load_runtime_paths_dict()
    return d["base_dir"], d["output_dir"], d["exception_file"]


BASE_DIR, OUTPUT_DIR, EXCEPTION_FILE = _load_runtime_paths()
#нормализация имен колонок
COLUMN_MAP = {
    "KUNNR": "Customer",
    "NAME1": "Name",
    "STCD": "Tax Number 1",
    "VKORD": "SO"
}

PAIRS = {
    "3801_3803": {"folders": ["3801", "3803"]},
    "3802_3804": {"folders": ["3802", "3804"]},
    "3805_3806": {"folders": ["3805", "3806"]},
}

# Как в Access (Joined_1): только <> "S" и <> "PR" для OrBlk и OrBlk1; TS не исключается.
BLOCKED_ORBLK = frozenset({"S", "PR"})
FAKE_INN = ("1111111111", "2222222222")

BILL_TO_COLS = [
    "SO",
    "Customer",
    "Name",
    "Tax Number 1",
    "CGrp",
    "Grp4",
    "Search Term 2",
    "SDst",
    "A7",
    "OrBlk1",
    "OrBlk2",
    "Cust OrBlk Bill-to",
    "BP now",
    "Check Cust&BP",
    "BP SO",
    "BP Customer",
    "BP Name",
    "BP Tax Number 1",
    "BP CGrp",
    "BP OrBlk1",
    "BP OrBlk2",
    "BP status",
]

# Лист несоответствий — как эталон Excel / Joined_1_Checks (порядок столбцов 1:1 со скрином).
MISMATCH_EXPORT_COLS = [
    "SO",
    "Customer",
    "Name",
    "Tax Number 1",
    "CGrp",
    "Grp4",
    "Search Term 2",
    "SDst",
    "A7",
    "OrBlk1",
    "OrBlk2",
    "Cust OrBlk Bill-to",
    "BP",
    "BP Name",
    "BP Tax Number 1",
    "BP OrBlk1",
    "BP OrBlk2",
    "BP OrBlk Bill-to",
    "PY",
    "PY Name",
    "PY Tax Number 1",
    "PY OrBlk1",
    "PY OrBlk2",
    "ZY",
    "ZY Name",
    "ZY Tax Number 1",
    "ZY OrBlk1",
    "ZY OrBlk2",
    "Check BP-PY-ZY",
    "Check Tax Number Cust&BP",
    "Check Cust&BP",
    "Comment MD Analyst",
    "Комментарий OM",
]

_CYR = re.compile(r"[А-Яа-яЁё]")


def _norm(v: object) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return ""
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def _norm_cust(v: object) -> str:
    s = _norm(v)
    if s.isdigit():
        return s.lstrip("0") or "0"
    return s


def _ns(s: pd.Series) -> pd.Series:
    return s.fillna("").astype(str).str.strip().apply(lambda x: x[:-2] if x.endswith(".0") else x)


def _nc(s: pd.Series) -> pd.Series:
    base = _ns(s)
    return base.where(~base.str.match(r"^\d+$"), base.str.lstrip("0").replace("", "0"))


def _align_merge_keys(df: pd.DataFrame) -> pd.DataFrame:
    """
    Приводаит Customer и _folder к согласованным строковым ключам.
    Иначе pandas.merge падает с int64 vs object — цепочка рвётся, отчёт без данных.
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    if "Customer" in out.columns:
        out["Customer"] = _nc(out["Customer"].fillna("").astype(str))
    if "_folder" in out.columns:
        out["_folder"] = out["_folder"].map(_norm)
    return out


def _col(df: pd.DataFrame, *candidates: str) -> str | None:
    lo = {c.casefold(): c for c in df.columns}
    for c in candidates:
        found = lo.get(c.casefold())
        if found is not None:
            return found
    return None


def _so_col(df: pd.DataFrame) -> str | None:
    return _col(df, "SOrg#", "SOrg.", "VKORG", "SO", "Sales Org.")


def _partner_lookup_extra_columns(base_lookup: pd.DataFrame, prefix: str) -> dict[str, str]:
    """
    Как в Access join к BP/PY/ZY: только поля master из Base для номера партнёра (без адресной «витрины»).
    """
    extra: dict[str, str] = {}
    logical_maps: list[tuple[tuple[str, ...], str]] = [
        (("Name",), f"{prefix} Name"),
        (("Tax Number 1", "STCD"), f"{prefix} Tax Number 1"),
        (("OrBlk",), f"{prefix} OrBlk1"),
        (("OrBlk1", "OrBlk 1"), f"{prefix} OrBlk2"),
    ]
    for candidates, dst in logical_maps:
        src = _col(base_lookup, *candidates)
        if src and src not in extra:
            extra[src] = dst
    return extra


def _get_file(folder: Path, pattern: str) -> Path | None:
    """Самый новый файл по дате изменения (если в папке несколько совпадений)."""
    if not folder.exists():
        return None
    files = [
        p for p in folder.glob(pattern)
        if not p.name.startswith("~$")  # lock-файл Excel при открытой книге — не xlsx
        and not p.name.startswith(".")
    ]
    if not files:
        return None
    if len(files) > 1:
        files = sorted(files, key=lambda p: p.stat().st_mtime, reverse=True)
        print(
            f"[build_checks] В {folder} найдено {len(files)} файлов «{pattern}», "
            f"берём новый: {files[0].name}",
            flush=True,
        )
    return files[0]


def _file_snapshot(path: Path) -> tuple[int, float]:
    st = path.stat()
    return st.st_size, st.st_mtime


def _peek_file_format(path: Path) -> str:
    """Сигнатура файла на диске (расширение .xlsx может врать)."""
    with open(path, "rb") as f:
        head = f.read(16)
    if head.startswith(b"PK"):
        return "xlsx_zip"
    if head.startswith(b"\xd0\xcf\x11\xe0"):
        return "xls_ole"
    stripped = head.lstrip()
    if stripped.startswith(b"<") or stripped.startswith(b"<?xml"):
        return "html_or_xml"
    return "unknown"


def _read_excel_robust(path: Path) -> pd.DataFrame:
    """
    Несколько движков read_excel.
    Excel часто открывает .xls под именем .xlsx — openpyxl падает с BadZipFile.
    """
    fmt = _peek_file_format(path)
    if fmt == "xlsx_zip":
        engines = ("openpyxl", "calamine")
    elif fmt == "xls_ole":
        engines = ("xlrd", "calamine")
    else:
        engines = ("openpyxl", "calamine", "xlrd")

    errors: list[str] = []
    for engine in engines:
        try:
            return pd.read_excel(path, dtype=str, engine=engine)
        except ImportError:
            errors.append(f"{engine}: не установлен (pip install {engine})")
        except Exception as exc:
            errors.append(f"{engine}: {type(exc).__name__}: {exc}")

    fmt_hint = {
        "xls_ole": (
            "На диске это старый Excel (.xls), хотя имя .xlsx. "
            "Excel открывает, openpyxl — нет. "
            "«Сохранить как» → xlsx или pip install xlrd"
        ),
        "html_or_xml": "Это HTML/XML, не Excel — перевыгрузите макросом.",
        "unknown": "Неизвестный формат файла.",
        "xlsx_zip": "ZIP/xlsx повреждён или нестандартный.",
    }.get(fmt, "")
    detail = "\n".join(f"  - {e}" for e in errors)
    if fmt_hint:
        detail += f"\n  {fmt_hint}"
    raise RuntimeError(f"Не удалось прочитать {path.name} ({fmt}):\n{detail}")


def read_excel_via_com(path: Path) -> pd.DataFrame:
    """
    Чтение через установленный Excel (Windows).
    Нужно, когда pandas/openpyxl пишут BadZipFile, а Excel файл открывает.
    """
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Чтение через Excel: pip install pywin32 (Windows + установленный Microsoft Excel)"
        ) from exc

    resolved = str(path.resolve())
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False
    try:
        wb = excel.Workbooks.Open(resolved, ReadOnly=True, UpdateLinks=0, CorruptLoad=1)
        try:
            ws = wb.Worksheets(1)
            used = ws.UsedRange
            nrows = int(used.Rows.Count)
            ncols = int(used.Columns.Count)
            if nrows < 1 or ncols < 1:
                return pd.DataFrame()
            raw = used.Value
        finally:
            wb.Close(False)
    finally:
        excel.Quit()

    if nrows == 1 and ncols == 1:
        rows: list[list[object]] = [[raw]]
    elif nrows == 1:
        rows = [list(raw)]
    elif ncols == 1:
        rows = [[v] for v in raw]
    else:
        rows = [list(r) for r in raw]

    header = ["" if c is None else str(c).strip() for c in rows[0]]
    header = _dedupe_column_labels(header)
    data = rows[1:]
    df = pd.DataFrame(data, columns=header)
    return df.astype(str).replace({"None": "", "nan": "", "<NA>": ""})


def _read_excel_with_com_fallback(path: Path, *, folder: str, kind: str) -> pd.DataFrame:
    try:
        return _read_excel_robust(path)
    except Exception as first_exc:
        if sys.platform != "win32":
            raise first_exc
        print(
            f"[build_checks] SO {folder} {kind}: pandas не прочитал ({type(first_exc).__name__}), "
            f"пробую через Excel…",
            flush=True,
        )
        return read_excel_via_com(path)


def _read_excel_checked(
    path: Path,
    *,
    kind: str,
    folder: str,
    allow_com: bool = False,
) -> pd.DataFrame:
    """read_excel с понятной ошибкой: какой файл и SO не читается."""
    snap_before = _file_snapshot(path)
    fmt = _peek_file_format(path)
    try:
        if allow_com:
            df = _read_excel_with_com_fallback(path, folder=folder, kind=kind)
        else:
            df = _read_excel_robust(path)
    except Exception as exc:
        if fmt == "xls_ole":
            hint = (
                "Файл открывается в Excel, но на диске это .xls (не .xlsx). "
                f"Размер: {snap_before[0]} байт. "
                "Сохраните в Excel как «Книга Excel (*.xlsx)» или: pip install xlrd"
            )
        elif fmt == "html_or_xml":
            hint = "Файл похож на HTML/XML, не на Excel — перевыгрузите макросом."
        elif "BadZipFile" in type(exc).__name__ or "zip" in str(exc).lower():
            hint = (
                f"BadZipFile (формат {fmt}). Excel иногда открывает после «ремонта». "
                f"Размер: {snap_before[0]} байт. "
                "Сохраните копию через Excel «Сохранить как» xlsx."
            )
        else:
            hint = str(exc)
        raise RuntimeError(
            f"Не удалось прочитать {kind} для SO {folder}:\n"
            f"  {path.resolve()}\n"
            f"  формат на диске: {fmt}\n"
            f"  {hint}\n"
            f"  Перевыгрузите файл макросом или удалите лишние/старые *{kind.split()[0]}* в папке."
        ) from exc
    snap_after = _file_snapshot(path)
    if snap_after != snap_before:
        raise RuntimeError(
            f"Файл изменился на диске во время чтения {kind} SO {folder}:\n"
            f"  {path.resolve()}\n"
            f"  до чтения: размер={snap_before[0]} mtime={snap_before[1]}\n"
            f"  после чтения: размер={snap_after[0]} mtime={snap_after[1]}\n"
            f"  Это не запись из pandas — ищите Excel, макрос выгрузки или синхронизацию (OneDrive)."
        )
    return df


def _dedupe_column_labels(columns) -> list[str]:
    """Повтор имени → Name.1, Name.2… (pandas и Excel COM с двумя OrBlk)."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for raw in columns:
        name = "" if raw is None else str(raw).strip()
        if name == "":
            name = "col"
        n = seen.get(name, 0)
        seen[name] = n + 1
        out.append(name if n == 0 else f"{name}.{n}")
    return out


def _normalize_orblk_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Внутренний вид Access: OrBlk (Bill-to = M) + OrBlk1 (второй блок).

    Макрос: два столбца OrBlk / OrBlk+OrBlk2 / OrBlk+OrBlk.1 — значения переставлены:
      2-й столбец файла = Access OrBlk (M)
      1-й столбец файла = Access OrBlk1

    Без swap M остаётся во 2-й колонке → десятки тысяч ложных
    «Ship-to прикреплён к BP без OB M» (типично 3802 с OrBlk.1).

    Уже Access (OrBlk+OrBlk1 без OrBlk2/OrBlk.1) — не трогаем.
    В отчёт: OrBlk→OrBlk1, OrBlk1→OrBlk2 (to_export_orblk_names).
    """
    if df is None or df.empty:
        return df

    out = df.copy()
    if len(out.columns) != len(set(map(str, out.columns))):
        out.columns = _dedupe_column_labels(out.columns)

    # OrBlk.1 (дубль заголовка) = второй столбец макроса = OrBlk2
    if "OrBlk.1" in out.columns and "OrBlk2" not in out.columns:
        out = out.rename(columns={"OrBlk.1": "OrBlk2"})
    elif "OrBlk.1" in out.columns and "OrBlk2" in out.columns:
        out = out.drop(columns=["OrBlk.1"])

    # Уже Access OrBlk+OrBlk1 — без макросного второго столбца
    if "OrBlk1" in out.columns and "OrBlk2" not in out.columns:
        return out

    if "OrBlk" not in out.columns:
        return out

    if "OrBlk2" in out.columns:
        return out.rename(columns={"OrBlk": "OrBlk1", "OrBlk2": "OrBlk"})

    return out


def to_export_orblk_names(df: pd.DataFrame) -> pd.DataFrame:
    """Access b.OrBlk AS OrBlk1, b.OrBlk1 AS OrBlk2 — внешний вид отчёта."""
    if df is None or df.empty:
        return df
    out = df.copy()
    if "OrBlk" in out.columns and "OrBlk1" in out.columns and "OrBlk2" not in out.columns:
        out = out.rename(columns={"OrBlk1": "OrBlk2", "OrBlk": "OrBlk1"})
    elif "OrBlk" in out.columns and "OrBlk1" not in out.columns:
        out = out.rename(columns={"OrBlk": "OrBlk1"})
    return out


def _normalize_base_df(df: pd.DataFrame, folder: str) -> pd.DataFrame:
    df = _normalize_orblk_columns(df)
    so = _so_col(df)
    if so and so != "SO":
        df = df.rename(columns={so: "SO"})
    cust = _col(df, "Customer", "KUNNR")
    if cust and cust != "Customer":
        df = df.rename(columns={cust: "Customer"})
    if "Customer" in df.columns:
        df["Customer"] = _nc(df["Customer"])
    for old, new in COLUMN_MAP.items():
        if old in df.columns and new not in df.columns:
            df = df.rename(columns={old: new})
    g4 = _col(df, "Grp4", "GROUP4", "Group4", "GRP4")
    if g4 and g4 != "Grp4":
        df = df.rename(columns={g4: "Grp4"})
    df["_folder"] = folder
    return df


def _read_base(path: Path, folder: str, *, allow_com: bool = False) -> pd.DataFrame:
    df = _read_excel_checked(path, kind="Base", folder=folder, allow_com=allow_com)
    return _normalize_base_df(df, folder)


def _normalize_partner_df(df: pd.DataFrame, folder: str, *, kind: str = "partner") -> pd.DataFrame:
    kunnr = _col(df, "KUNNR", "Customer", "Sold-to")
    ktonr = _col(df, "KTONR", "BP", "PY", "ZY")
    if not kunnr or not ktonr:
        return pd.DataFrame(columns=["KUNNR", "KTONR", "_folder"])
    out = df[[kunnr, ktonr]].rename(columns={kunnr: "KUNNR", ktonr: "KTONR"})
    out["KUNNR"] = _nc(out["KUNNR"])
    out["KTONR"] = _nc(out["KTONR"])
    out["_folder"] = folder
    return out.dropna(subset=["KUNNR"]).drop_duplicates(subset=["KUNNR", "_folder"], keep="first")


def _read_partner(
    path: Path,
    folder: str,
    *,
    kind: str = "partner",
    allow_com: bool = False,
) -> pd.DataFrame:
    df = _read_excel_checked(path, kind=kind, folder=folder, allow_com=allow_com)
    return _normalize_partner_df(df, folder, kind=kind)


def load_exception(base_dir: Path) -> pd.DataFrame:
    parts = []
    for f in base_dir.rglob("Exception.xlsx"):
        try:
            parts.append(pd.read_excel(f, dtype=str))
        except Exception:
            pass
    if EXCEPTION_FILE.exists():
        try:
            parts.append(pd.read_excel(EXCEPTION_FILE, dtype=str))
        except Exception:
            pass
    if not parts:
        return pd.DataFrame(columns=["SO", "Customer", "Comment OM"])
    df = pd.concat(parts, ignore_index=True)
    so_c = _col(df, "SO", "SOrg#", "SOrg.", "VKORG")
    cu_c = _col(df, "Customer", "KUNNR")
    if not so_c or not cu_c:
        return pd.DataFrame(columns=["SO", "Customer", "Comment OM"])
    co_c = _col(df, "Comment OM", "Comment", "Комментарий")
    cols: dict[str, str] = {so_c: "SO", cu_c: "Customer"}
    if co_c:
        cols[co_c] = "Comment OM"
    out = df.rename(columns=cols)[list(cols.values())].copy()
    if "Comment OM" not in out.columns:
        out["Comment OM"] = ""
    out["SO"] = _ns(out["SO"])
    out["Customer"] = _nc(out["Customer"])
    out = out[(out["SO"] != "") & (out["Customer"] != "")]
    return out.drop_duplicates(subset=["SO", "Customer"], keep="last").reset_index(drop=True)


def exception_keys(exc_df: pd.DataFrame) -> Set[Tuple[str, str]]:
    if exc_df.empty:
        return set()
    return {(_norm(r["SO"]), _norm_cust(r["Customer"])) for _, r in exc_df.iterrows()}


def collect_and_persist_global_exception(base_dir: Path, output_dir: Path) -> pd.DataFrame:
    exc = load_exception(base_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    if exc.empty:
        try:
            (output_dir / "00_Exception_собрано_нет_данных.txt").write_text(
                "Исключения не загружены\n", encoding="utf-8")
        except Exception:
            pass
    else:
        try:
            exc.to_excel(output_dir / "00_Exception_собрано_перед_отчётами.xlsx",
                         index=False, sheet_name="Exception")
        except Exception:
            pass
    return exc


def _prep_bill_to_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Шапка листа «Привязка Bill-to по ИНН» — q_*_Joined_3_BillToByINN."""
    if df is None or df.empty:
        return pd.DataFrame(columns=BILL_TO_COLS)

    out = df.copy()
    if "Grp4" not in out.columns:
        g4 = _col(out, "Grp4", "GROUP4", "Group4", "GRP4")
        if g4:
            out = out.rename(columns={g4: "Grp4"})
    if "Grp4" not in out.columns:
        out["Grp4"] = ""

    if "BP now" not in out.columns and "BP" in out.columns:
        out = out.rename(columns={"BP": "BP now"})

    if "_folder" in out.columns:
        so_series = _ns(out["SO"]) if "SO" in out.columns else pd.Series("", index=out.index, dtype=object)
        need = so_series.eq("") | so_series.isna()
        if need.any():
            if "SO" not in out.columns:
                out["SO"] = ""
            out.loc[need, "SO"] = _ns(out.loc[need, "_folder"])

    for col in BILL_TO_COLS:
        if col not in out.columns:
            out[col] = ""
    return out[BILL_TO_COLS]


def _excel_sheet_name_safe(name: str) -> str:
    """Имя листа Excel ≤31 символа, без : \\ / ? * [ ]"""
    s = re.sub(r'[:\\/?*\[\]]', " ", str(name)).strip()
    return (s[:31] if s else "Sheet").strip() or "Sheet"


def _unique_excel_sheet_name(proposed: str, used: set[str]) -> str:
    base = _excel_sheet_name_safe(proposed)
    if base not in used:
        return base
    for i in range(2, 999):
        suf = f" {i}"
        cand = _excel_sheet_name_safe(base[: max(1, 31 - len(suf))] + suf)
        if cand not in used:
            return cand
    return _excel_sheet_name_safe(base[:28] + "...")[:31]


def _attach_comment_om(df: pd.DataFrame, exc_df: pd.DataFrame) -> pd.DataFrame:
    """Подмешать Comment OM из Exception по нормализованным SO + Customer."""
    if df.empty:
        return df
    if exc_df is None or exc_df.empty:
        out = df.copy()
        if "Comment OM" not in out.columns:
            out["Comment OM"] = ""
        return out

    left = df.copy()
    if "Comment OM" in left.columns:
        left = left.drop(columns=["Comment OM"])

    ex = exc_df[["SO", "Customer", "Comment OM"]].copy()
    ex["_j_so"] = _ns(ex["SO"])
    ex["_j_cu"] = _nc(ex["Customer"])
    ex = ex.drop_duplicates(subset=["_j_so", "_j_cu"], keep="last")[["_j_so", "_j_cu", "Comment OM"]]

    left["_j_so"] = _ns(left["SO"]) if "SO" in left.columns else ""
    left["_j_cu"] = _nc(left["Customer"])
    merged = left.merge(ex, on=["_j_so", "_j_cu"], how="left")
    merged = merged.drop(columns=["_j_so", "_j_cu"])
    merged["Comment OM"] = merged["Comment OM"].fillna("")
    return merged


def _dedupe_mismatch_report_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Одна строка отчёта на комбинацию (орг. выгрузка, ИНН, BP, текст Comment MD Analyst).
    Одинаковая проблема по многим ship-to не раздувает список; одна и та же проблема в 3801 и 3803 —
    остаётся две строки (разные _folder).
    Оставляем первую строку после сортировки по SO, Customer (стабильный «представитель»).
    """
    if df.empty:
        return df
    if not all(k in df.columns for k in ("Tax Number 1", "BP", "Comment MD Analyst")):
        return df
    work = df.copy()
    sort_cols = [c for c in ("SO", "Customer") if c in work.columns]
    if sort_cols:
        work = work.sort_values(by=sort_cols, kind="stable")
    work["_d_tax"] = work["Tax Number 1"].fillna("").astype(str).str.strip()
    work["_d_bp"] = work["BP"].fillna("").astype(str).map(_norm_cust)
    work["_d_cm"] = work["Comment MD Analyst"].fillna("").astype(str).str.strip()
    dedupe_cols = ["_d_tax", "_d_bp", "_d_cm"]
    if "_folder" in work.columns:
        work["_d_fd"] = work["_folder"].fillna("").astype(str).map(_norm)
        dedupe_cols.insert(0, "_d_fd")
    out = work.drop_duplicates(subset=dedupe_cols, keep="first").reset_index(drop=True)
    return out.drop(columns=[c for c in ("_d_tax", "_d_bp", "_d_cm", "_d_fd") if c in out.columns])


def _mismatch_export_value(row: pd.Series, export_col: str) -> str:
    """Значение для колонки экспорта; имена в Excel могут отличаться от внутренних (checked)."""
    src_cols: tuple[str, ...]
    if export_col == "Check Tax Number Cust&BP":
        src_cols = ("Check Tax Number Cust&BP", "Check Tax Number Cust-BP")
    elif export_col == "Комментарий OM":
        src_cols = ("Комментарий OM", "Comment OM", "Comment")
    else:
        src_cols = (export_col,)

    for src in src_cols:
        if src not in row.index:
            continue
        v = row[src]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if s.lower() in ("nan", "<na>", "none"):
            continue
        return s
    return ""


def _prep_mismatch_sheet(df: pd.DataFrame) -> pd.DataFrame:
    """Строки = результат add_checks (filtered); шапка как Joined_1_Checks, не витрина адресов."""
    if df is None or df.empty:
        return pd.DataFrame(columns=MISMATCH_EXPORT_COLS)

    out = df.copy()
    if "Grp4" not in out.columns:
        g4 = _col(out, "Grp4", "GROUP4", "Group4", "GRP4")
        if g4:
            out = out.rename(columns={g4: "Grp4"})
    if "Grp4" not in out.columns:
        out["Grp4"] = ""

    if "_folder" in out.columns:
        so_series = _ns(out["SO"]) if "SO" in out.columns else pd.Series("", index=out.index, dtype=object)
        need = so_series.eq("") | so_series.isna()
        if need.any():
            if "SO" not in out.columns:
                out["SO"] = ""
            out.loc[need, "SO"] = _ns(out.loc[need, "_folder"])

    rows: list[dict[str, str]] = []
    for _, row in out.iterrows():
        rows.append({c: _mismatch_export_value(row, c) for c in MISMATCH_EXPORT_COLS})

    return pd.DataFrame(rows, columns=MISMATCH_EXPORT_COLS)


# Заливка листов Excel (тема Excel).
_FILL_OLIVE_60 = "C4D79B"       # Olive Green, Lighter 60%
_FILL_BLUE_80 = "DCE6F1"        # Blue, Lighter 80%
_FILL_ORANGE_80 = "FDE9D9"      # Orange, Lighter 80%
_FILL_PURPLE_80 = "E4DFEC"      # Purple, Lighter 80%
_FILL_AQUA_80 = "DAEEF3"        # Aqua, Lighter 80%
_FILL_YELLOW = "FFFF00"
_FILL_WHITE_DARK_15 = "D9D9D9"  # White, Dark 15%

# Лист «… Несоответствия» — заливка всей колонки.
_MISMATCH_COLUMN_FILL_HEX: dict[str, str] = {
    "Customer": _FILL_OLIVE_60,
    "OrBlk1": _FILL_OLIVE_60,
    "OrBlk2": _FILL_OLIVE_60,
    "Cust OrBlk Bill-to": _FILL_OLIVE_60,
    "Check BP-PY-ZY": _FILL_OLIVE_60,
    "Check Tax Number Cust&BP": _FILL_OLIVE_60,
    "Check Cust&BP": _FILL_OLIVE_60,
    "BP": _FILL_BLUE_80,
    "BP OrBlk1": _FILL_BLUE_80,
    "BP OrBlk2": _FILL_BLUE_80,
    "BP OrBlk Bill-to": _FILL_BLUE_80,
    "PY": _FILL_ORANGE_80,
    "PY OrBlk1": _FILL_ORANGE_80,
    "PY OrBlk2": _FILL_ORANGE_80,
    "ZY": _FILL_PURPLE_80,
    "ZY OrBlk1": _FILL_PURPLE_80,
    "ZY OrBlk2": _FILL_PURPLE_80,
    "Comment MD Analyst": _FILL_YELLOW,
}

# Лист «… Привязка Bill-to по ИНН» — заливка всей колонки.
_BILL_TO_COLUMN_FILL_HEX: dict[str, str] = {
    "Customer": _FILL_OLIVE_60,
    "OrBlk1": _FILL_OLIVE_60,
    "OrBlk2": _FILL_OLIVE_60,
    "Cust OrBlk Bill-to": _FILL_OLIVE_60,
    "BP now": _FILL_OLIVE_60,
    "Check Cust&BP": _FILL_OLIVE_60,
    "BP Customer": _FILL_AQUA_80,
    "BP OrBlk1": _FILL_AQUA_80,
    "BP OrBlk2": _FILL_AQUA_80,
    "BP status": _FILL_YELLOW,
}


def _excel_cell_display_width(value: object) -> float:
    if value is None:
        return 0.0
    s = str(value)
    return sum(1.6 if ord(ch) > 127 else 1.0 for ch in s)


def _autofit_worksheet_columns(ws, *, min_width: float = 6.0, max_width: float = 72.0) -> None:
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, (ws.max_column or 0) + 1):
        letter = get_column_letter(col_idx)
        max_len = 0.0
        for row in range(1, (ws.max_row or 0) + 1):
            max_len = max(max_len, _excel_cell_display_width(ws.cell(row=row, column=col_idx).value))
        ws.column_dimensions[letter].width = min(max(max_len + 1.5, min_width), max_width)


def _format_worksheet_columns(
    ws,
    full_column_hex: dict[str, str],
    *,
    header_only_rest_hex: str | None = _FILL_WHITE_DARK_15,
) -> None:
    """Указанные колонки — заливка целиком; остальные — только шапка (если задан цвет)."""
    from openpyxl.styles import PatternFill

    if not ws.max_column or not ws.max_row:
        return
    if ws.max_column == 1 and str(ws.cell(1, 1).value or "").strip() == "Сообщение":
        _autofit_worksheet_columns(ws)
        return

    full_fills = {
        name: PatternFill(fill_type="solid", fgColor=hex_color)
        for name, hex_color in full_column_hex.items()
    }
    header_rest = (
        PatternFill(fill_type="solid", fgColor=header_only_rest_hex)
        if header_only_rest_hex
        else None
    )

    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(1, col_idx).value or "").strip()
        if not header:
            continue
        full_fill = full_fills.get(header)
        if full_fill is not None:
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).fill = full_fill
        elif header_rest is not None:
            ws.cell(row=1, column=col_idx).fill = header_rest

    _autofit_worksheet_columns(ws)


def _format_mismatch_worksheet(ws) -> None:
    """Лист «… Несоответствия»."""
    _format_worksheet_columns(ws, _MISMATCH_COLUMN_FILL_HEX)


def _format_bill_to_worksheet(ws) -> None:
    """Лист «… Привязка Bill-to по ИНН»."""
    _format_worksheet_columns(ws, _BILL_TO_COLUMN_FILL_HEX)


def _format_exception_worksheet(ws) -> None:
    """Лист Exception — жёлтая шапка."""
    from openpyxl.styles import PatternFill

    if not ws.max_column or not ws.max_row:
        return

    yellow = PatternFill(fill_type="solid", fgColor=_FILL_YELLOW)
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col_idx).fill = yellow
    _autofit_worksheet_columns(ws)


def _write_mismatch_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    safe = _excel_sheet_name_safe(sheet_name)
    df.to_excel(writer, sheet_name=safe, index=False)
    _format_mismatch_worksheet(writer.sheets[safe])


def _write_bill_to_sheet(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    safe = _excel_sheet_name_safe(sheet_name)
    df.to_excel(writer, sheet_name=safe, index=False)
    _format_bill_to_worksheet(writer.sheets[safe])


def _write_exception_sheet(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    safe = _excel_sheet_name_safe("Exception")
    df.to_excel(writer, sheet_name=safe, index=False)
    _format_exception_worksheet(writer.sheets[safe])


def _mismatch_sheet_label(so_token: str) -> str:
    """Как у Bill-to: префикс SOrg + тема листа."""
    so = _norm(so_token)
    if not so:
        so = "SO"
    return f"{so} Несоответствия"
